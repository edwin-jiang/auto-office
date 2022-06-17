import time
import PySimpleGUI as sg
from openpyxl import Workbook,load_workbook
import pandas as pd

#操作区分
optype = 0
# 主题色
sg.theme('DarkAmber')
# 窗口布局
layout = [
    [sg.Text('文档操作小工具 ver0.1')],
    [sg.Multiline('', size=(100,10), key='textContent')],
    [sg.Text('操作选项')],
    [sg.Radio('创建Excel', 'S1', enable_events=True, key='id0', default=True),
     sg.Radio('读取Excel', 'S1', enable_events=True, key='id1'),],
    [sg.Button('执行', key='execBtn'),]
]

def createExcelFile():
    # 创建一个工作簿对象
    wb = Workbook()
    # 在索引为0的位置创建一个名为mytest的sheet页
    ws = wb.create_sheet('mytest',0)
    # 对sheet页设置一个颜色（16位的RGB颜色）
    ws.sheet_properties.tabColor = 'ff72BA'
    # 将创建的工作簿保存为Mytest.xlsx
    wb.save('Mytest.xlsx')
    # 最后关闭文件
    wb.close()

def openExcelFile(values):
    filename = 'Mytest.xlsx'
    # 加载工作簿
    wb = load_workbook(filename)
    # 获取sheet页
    ws = wb['mytest']
    ws.cell(row=1,column=1).value = values
    # 打印sheet页的颜色属性值
    print('color:',ws.sheet_properties.tabColor)
    wb.save(filename)
    wb.close()

# 创建窗口
window = sg.Window('文档操作小工具 ver0.1', layout)

# 循环处理事件
while True:
    event, values = window.read()

    # 用户点击X关闭窗口或点击退出按钮
    if event == sg.WIN_CLOSED:
        break
    if event == 'id0':
        optype = 0
    if event == 'id1':
        optype = 1
    print(optype)

    if event == 'execBtn':
        if optype == 0:
            createExcelFile()
        else:
            openExcelFile(values['textContent'])

window.close()